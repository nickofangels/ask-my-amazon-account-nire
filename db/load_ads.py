"""
db/load_ads.py — Load Amazon Advertising Console exports into Supabase.

Reads raw/ads/*.xlsx and raw/ads/*.csv, normalizes column names, and
upserts into ads_* tables. Separate from the organic data pipeline.

Usage:
    python -m db.load_ads                   # load everything
    python -m db.load_ads --only campaigns  # load one table
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import psycopg2.extras

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from config import ADS_DIR
from schema import get_conn, init_db


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _upsert(conn, sql: str, rows: list[tuple], page_size: int = 500) -> None:
    if not rows:
        return
    with conn.cursor() as cur:
        psycopg2.extras.execute_values(cur, sql, rows, page_size=page_size)
    conn.commit()


def _log_pull(conn, report_type: str, month: str | None, source_file: str,
              row_count: int, now: str) -> None:
    with conn.cursor() as cur:
        cur.execute(
            "INSERT INTO pull_log (report_type, month, source_file, row_count, pulled_at) "
            "VALUES (%s, %s, %s, %s, %s)",
            (report_type, month, source_file, row_count, now),
        )
    conn.commit()


def _read_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else '' for h in rows[0]]
    return [{h: v for h, v in zip(headers, row)} for row in rows[1:]]


def _read_csv(path: Path) -> list[dict]:
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        return [{k.strip(): v for k, v in row.items()} for row in reader]


def _read_report(path: Path) -> list[dict]:
    if path.suffix == '.csv':
        return _read_csv(path)
    return _read_xlsx(path)


def _clean_num(val) -> float | None:
    """Convert '$59.61', '12.246%', '100.0%' etc to float. Returns None for empty/nan."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ('nan', 'none', '-', ''):
        return None
    s = s.replace('$', '').replace(',', '').replace('%', '')
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _clean_int(val) -> int | None:
    n = _clean_num(val)
    return int(n) if n is not None else None


def _extract_month(row: dict) -> str:
    """Extract YYYY-MM from Start Date column."""
    from datetime import datetime as _dt
    raw = row.get('Start Date') or row.get('start_date') or ''
    if isinstance(raw, _dt):
        return raw.strftime('%Y-%m')
    s = str(raw).strip()
    if not s:
        return ''
    # Try ISO format first (2026-02-01)
    for fmt in ('%Y-%m-%d', '%b %d, %Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return _dt.strptime(s.split()[0] if 'T' not in s else s[:10], fmt[:len(fmt)]).strftime('%Y-%m')
        except (ValueError, IndexError):
            continue
    # Fallback: pandas
    try:
        import pandas as pd
        return pd.to_datetime(s).strftime('%Y-%m')
    except Exception:
        return ''


def _extract_month_batch(rows: list[dict]) -> str:
    """Get the most common month from a batch of rows."""
    if not rows:
        return ''
    return _extract_month(rows[0])


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_ads_campaigns(conn) -> None:
    """SP Campaign + SP Budget + SB Campaign + SD Campaign → ads_campaigns."""
    now = _now()
    sql = """
        INSERT INTO ads_campaigns VALUES %s
        ON CONFLICT (month, ad_type, campaign_name) DO UPDATE SET
            attribution_window=EXCLUDED.attribution_window, status=EXCLUDED.status,
            portfolio_name=EXCLUDED.portfolio_name, budget=EXCLUDED.budget,
            spend=EXCLUDED.spend, impressions=EXCLUDED.impressions,
            clicks=EXCLUDED.clicks, ctr=EXCLUDED.ctr, cpc=EXCLUDED.cpc,
            orders=EXCLUDED.orders, units=EXCLUDED.units, sales=EXCLUDED.sales,
            acos=EXCLUDED.acos, roas=EXCLUDED.roas, cvr=EXCLUDED.cvr,
            bidding_strategy=EXCLUDED.bidding_strategy,
            targeting_type=EXCLUDED.targeting_type,
            recommended_budget=EXCLUDED.recommended_budget,
            avg_time_in_budget=EXCLUDED.avg_time_in_budget,
            est_missed_imp_lower=EXCLUDED.est_missed_imp_lower,
            est_missed_imp_upper=EXCLUDED.est_missed_imp_upper,
            est_missed_clicks_lower=EXCLUDED.est_missed_clicks_lower,
            est_missed_clicks_upper=EXCLUDED.est_missed_clicks_upper,
            est_missed_sales_lower=EXCLUDED.est_missed_sales_lower,
            est_missed_sales_upper=EXCLUDED.est_missed_sales_upper,
            ntb_orders=EXCLUDED.ntb_orders, ntb_sales=EXCLUDED.ntb_sales,
            ntb_units=EXCLUDED.ntb_units, ntb_order_pct=EXCLUDED.ntb_order_pct,
            ntb_sales_pct=EXCLUDED.ntb_sales_pct,
            branded_searches=EXCLUDED.branded_searches, dpv=EXCLUDED.dpv,
            video_complete_views=EXCLUDED.video_complete_views,
            video_completion_rate=EXCLUDED.video_completion_rate,
            pulled_at=EXCLUDED.pulled_at
    """

    all_rows = []

    # --- SP Campaign + SP Budget (merge on campaign_name) ---
    sp_camp = _read_report(ADS_DIR / 'sp_campaign.csv')
    sp_budget_raw = _read_report(ADS_DIR / 'sp_budget.csv')
    budget_by_camp = {}
    for r in sp_budget_raw:
        cn = (r.get('Campaign Name') or '').strip()
        if cn:
            budget_by_camp[cn] = r

    for r in sp_camp:
        month = _extract_month(r)
        cn = (r.get('Campaign Name') or '').strip()
        b = budget_by_camp.get(cn, {})
        all_rows.append((
            month, 'SP', cn, '7d',
            (r.get('Status') or '').strip(),
            (r.get('Portfolio name') or '').strip(),
            _clean_num(r.get('Budget Amount')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_int(r.get('7 Day Total Orders (#)')),
            None,  # units not in SP campaign report
            _clean_num(r.get('7 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            None,  # cvr
            (r.get('Bidding strategy') or '').strip() or None,
            (r.get('Targeting Type') or '').strip() or None,
            _clean_num(b.get('Recommended Budget')),
            _clean_num(b.get('Average Time in Budget')),
            _clean_num(b.get('Estimated Missed Impressions Range Min')),
            _clean_num(b.get('Estimated Missed Impressions Range Max')),
            _clean_num(b.get('Estimated Missed Clicks Range Min')),
            _clean_num(b.get('Estimated Missed Clicks Range Max')),
            _clean_num(b.get('Estimated Missed Sales Range Min')),
            _clean_num(b.get('Estimated Missed Sales Range Max')),
            None, None, None, None, None,  # ntb_orders, ntb_sales, ntb_units, ntb_order_pct, ntb_sales_pct
            None, None,  # branded_searches, dpv
            None, None,  # video_complete_views, video_completion_rate
            now,
        ))

    # --- SB Campaign ---
    sb_camp = _read_report(ADS_DIR / 'sb_campaign.xlsx')
    for r in sb_camp:
        month = _extract_month(r)
        all_rows.append((
            month, 'SB', (r.get('Campaign Name') or '').strip(), '14d',
            None,
            (r.get('Portfolio name') or '').strip(),
            None,  # budget
            _clean_num(r.get('Spend')),
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_int(r.get('14 Day Total Orders (#)')),
            _clean_int(r.get('14 Day Total Units (#)')),
            _clean_num(r.get('14 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            _clean_num(r.get('14 Day Conversion Rate')),
            None, None,  # bidding, targeting
            None, None,  # budget reco
            None, None, None, None, None, None,  # missed
            _clean_int(r.get('14 Day New-to-brand Orders (#)')),
            _clean_num(r.get('14 Day New-to-brand Sales')),
            _clean_int(r.get('14 Day New-to-brand Units (#)')),
            _clean_num(r.get('14 Day % of Orders New-to-brand')),
            _clean_num(r.get('14 Day % of Sales New-to-brand')),
            _clean_int(r.get('14 Day Branded Searches')),
            _clean_int(r.get('14 Day Detail Page Views (DPV)')),
            _clean_int(r.get('Video Complete Views')),
            _clean_num(r.get('5 Second View Rate')),
            now,
        ))

    # --- SD Campaign ---
    sd_camp = _read_report(ADS_DIR / 'sd_campaign.xlsx')
    for r in sd_camp:
        month = _extract_month(r)
        all_rows.append((
            month, 'SD', (r.get('Campaign Name') or '').strip(), '14d',
            (r.get('Status') or '').strip(),
            (r.get('Portfolio name') or '').strip(),
            _clean_num(r.get('Budget Amount')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_int(r.get('14 Day Total Orders (#)')),
            _clean_int(r.get('14 Day Total Units (#)')),
            _clean_num(r.get('14 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            None,  # cvr
            None, None,  # bidding, targeting
            None, None,  # budget reco
            None, None, None, None, None, None,  # missed
            _clean_int(r.get('14 Day New-to-brand Orders (#)')),
            _clean_num(r.get('14 Day New-to-brand Sales')),
            _clean_int(r.get('14 Day New-to-brand Units (#)')),
            _clean_num(r.get('14 Day % of Orders New-to-brand')),
            _clean_num(r.get('14 Day % of Sales New-to-brand')),
            _clean_int(r.get('14 Day Branded Searches')),
            _clean_int(r.get('14 Day Detail Page Views (DPV)')),
            None, None,  # video
            now,
        ))

    _upsert(conn, sql, all_rows)
    month = _extract_month_batch(sp_camp or sb_camp or sd_camp)
    print(f"  [ads_campaigns] {len(all_rows)} rows")
    _log_pull(conn, 'ads_campaigns', month, 'sp+sb+sd_campaign', len(all_rows), now)


def load_ads_search_terms(conn) -> None:
    """SP + SB search term + impression share → ads_search_terms."""
    now = _now()
    sql = """
        INSERT INTO ads_search_terms VALUES %s
        ON CONFLICT (month, ad_type, campaign_name, ad_group_name,
                     targeting_text, customer_search_term) DO UPDATE SET
            match_type=EXCLUDED.match_type, attribution_window=EXCLUDED.attribution_window,
            impressions=EXCLUDED.impressions, clicks=EXCLUDED.clicks, ctr=EXCLUDED.ctr,
            cpc=EXCLUDED.cpc, spend=EXCLUDED.spend, orders=EXCLUDED.orders,
            units=EXCLUDED.units, sales=EXCLUDED.sales, acos=EXCLUDED.acos,
            roas=EXCLUDED.roas, cvr=EXCLUDED.cvr,
            impression_rank=EXCLUDED.impression_rank,
            impression_share=EXCLUDED.impression_share,
            own_sku_units=EXCLUDED.own_sku_units, own_sku_sales=EXCLUDED.own_sku_sales,
            other_sku_units=EXCLUDED.other_sku_units, other_sku_sales=EXCLUDED.other_sku_sales,
            pulled_at=EXCLUDED.pulled_at
    """
    all_rows = []

    # --- SP Search Term + SP Impression Share ---
    sp_st = _read_report(ADS_DIR / 'sp_search_term.xlsx')
    sp_is = _read_report(ADS_DIR / 'sp_search_term_impression_share.csv')

    # Index impression share by (campaign, ad_group, targeting, search_term)
    imp_share_idx: dict[tuple, dict] = {}
    for r in sp_is:
        key = (
            (r.get('Campaign Name') or '').strip(),
            (r.get('Ad Group Name') or '').strip(),
            (r.get('Targeting') or '').strip(),
            (r.get('Customer Search Term') or '').strip().lower(),
        )
        imp_share_idx[key] = r

    for r in sp_st:
        month = _extract_month(r)
        cn = (r.get('Campaign Name') or '').strip()
        ag = (r.get('Ad Group Name') or '').strip()
        tgt = (r.get('Targeting') or '').strip()
        cst = (r.get('Customer Search Term') or '').strip()
        ish = imp_share_idx.get((cn, ag, tgt, cst.lower()), {})
        all_rows.append((
            month, 'SP', cn, ag, tgt,
            (r.get('Match Type') or '').strip(),
            cst, '7d',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('7 Day Total Orders (#)')),
            _clean_int(r.get('7 Day Total Units (#)')),
            _clean_num(r.get('7 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            _clean_num(r.get('7 Day Conversion Rate')),
            _clean_int(ish.get('Search Term Impression Rank')),
            _clean_num(ish.get('Search Term Impression Share')),
            _clean_int(r.get('7 Day Advertised SKU Units (#)')),
            _clean_num(r.get('7 Day Advertised SKU Sales')),
            _clean_int(r.get('7 Day Other SKU Units (#)')),
            _clean_num(r.get('7 Day Other SKU Sales')),
            now,
        ))

    # --- SB Search Term + SB Impression Share ---
    sb_st = _read_report(ADS_DIR / 'sb_search_term.xlsx')
    sb_is = _read_report(ADS_DIR / 'sb_search_term_impression_share.csv')
    sb_ish_idx: dict[tuple, dict] = {}
    for r in sb_is:
        key = (
            (r.get('Campaign Name') or '').strip(),
            (r.get('Ad Group Name') or '').strip(),
            (r.get('Targeting') or '').strip(),
            (r.get('Customer Search Term') or '').strip().lower(),
        )
        sb_ish_idx[key] = r

    for r in sb_st:
        month = _extract_month(r)
        cn = (r.get('Campaign Name') or '').strip()
        ag = (r.get('Ad Group Name') or '').strip()
        tgt = (r.get('Targeting') or '').strip()
        cst = (r.get('Customer Search Term') or '').strip()
        ish = sb_ish_idx.get((cn, ag, tgt, cst.lower()), {})
        all_rows.append((
            month, 'SB', cn, ag, tgt,
            (r.get('Match Type') or '').strip(),
            cst, '14d',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('14 Day Total Orders (#)')),
            _clean_int(r.get('14 Day Total Units (#)')),
            _clean_num(r.get('14 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            _clean_num(r.get('14 Day Conversion Rate')),
            _clean_int(ish.get('Search Term Impression Rank')),
            _clean_num(ish.get('Search Term Impression Share')),
            None, None, None, None,  # own/other sku (not in SB)
            now,
        ))

    _upsert(conn, sql, all_rows)
    month = _extract_month_batch(sp_st or sb_st)
    print(f"  [ads_search_terms] {len(all_rows)} rows")
    _log_pull(conn, 'ads_search_terms', month, 'sp+sb_search_term', len(all_rows), now)


def load_ads_targeting(conn) -> None:
    """SP Targeting + SB Keyword + SD Targeting + SP Audience + SD Matched Target."""
    now = _now()
    sql = """
        INSERT INTO ads_targeting VALUES %s
        ON CONFLICT (month, ad_type, campaign_name, ad_group_name,
                     targeting_text, match_type) DO UPDATE SET
            attribution_window=EXCLUDED.attribution_window,
            targeting_type=EXCLUDED.targeting_type,
            impressions=EXCLUDED.impressions, clicks=EXCLUDED.clicks,
            ctr=EXCLUDED.ctr, cpc=EXCLUDED.cpc, spend=EXCLUDED.spend,
            orders=EXCLUDED.orders, units=EXCLUDED.units, sales=EXCLUDED.sales,
            acos=EXCLUDED.acos, roas=EXCLUDED.roas, cvr=EXCLUDED.cvr,
            top_of_search_imp_share=EXCLUDED.top_of_search_imp_share,
            own_sku_units=EXCLUDED.own_sku_units, own_sku_sales=EXCLUDED.own_sku_sales,
            other_sku_units=EXCLUDED.other_sku_units, other_sku_sales=EXCLUDED.other_sku_sales,
            ntb_orders=EXCLUDED.ntb_orders, ntb_sales=EXCLUDED.ntb_sales,
            branded_searches=EXCLUDED.branded_searches,
            pulled_at=EXCLUDED.pulled_at
    """
    all_rows = []

    # SP Targeting
    for r in _read_report(ADS_DIR / 'sp_targeting.xlsx'):
        month = _extract_month(r)
        all_rows.append((
            month, 'SP',
            (r.get('Campaign Name') or '').strip(),
            (r.get('Ad Group Name') or '').strip(),
            (r.get('Targeting') or '').strip(),
            (r.get('Match Type') or '').strip(),
            '7d', 'keyword',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('7 Day Total Orders (#)')),
            _clean_int(r.get('7 Day Total Units (#)')),
            _clean_num(r.get('7 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            _clean_num(r.get('7 Day Conversion Rate')),
            _clean_num(r.get('Top-of-search Impression Share')),
            _clean_int(r.get('7 Day Advertised SKU Units (#)')),
            _clean_num(r.get('7 Day Advertised SKU Sales')),
            _clean_int(r.get('7 Day Other SKU Units (#)')),
            _clean_num(r.get('7 Day Other SKU Sales')),
            None, None, None,  # ntb, branded
            now,
        ))

    # SB Keyword
    for r in _read_report(ADS_DIR / 'sb_keyword.xlsx'):
        month = _extract_month(r)
        all_rows.append((
            month, 'SB',
            (r.get('Campaign Name') or '').strip(),
            (r.get('Ad Group Name') or '').strip(),
            (r.get('Targeting') or '').strip(),
            (r.get('Match Type') or '').strip(),
            '14d', 'keyword',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('14 Day Total Orders (#)')),
            _clean_int(r.get('14 Day Total Units (#)')),
            _clean_num(r.get('14 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            _clean_num(r.get('14 Day Conversion Rate')),
            _clean_num(r.get('Top-of-search Impression Share')),
            None, None, None, None,  # own/other sku
            _clean_int(r.get('14 Day New-to-brand Orders (#)')),
            _clean_num(r.get('14 Day New-to-brand Sales')),
            _clean_int(r.get('14 Day Branded Searches')),
            now,
        ))

    # SD Targeting
    for r in _read_report(ADS_DIR / 'sd_targeting.xlsx'):
        month = _extract_month(r)
        all_rows.append((
            month, 'SD',
            (r.get('Campaign Name') or '').strip(),
            (r.get('Ad Group Name') or '').strip(),
            (r.get('Targeting') or '').strip(),
            '',  # no match type for SD
            '14d', 'audience',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('14 Day Total Orders (#)')),
            _clean_int(r.get('14 Day Total Units (#)')),
            _clean_num(r.get('14 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            None,  # cvr
            None,  # top-of-search
            None, None, None, None,  # own/other sku
            _clean_int(r.get('14 Day New-to-brand Orders (#)')),
            _clean_num(r.get('14 Day New-to-brand Sales')),
            None,  # branded
            now,
        ))

    # SP Audience (small, targeting_type='audience')
    for r in _read_report(ADS_DIR / 'sp_audience.xlsx'):
        month = _extract_month(r)
        all_rows.append((
            month, 'SP',
            (r.get('Campaign Name') or '').strip(),
            '',  # no ad group
            (r.get('Audience Name') or '').strip(),
            '',  # no match type
            '7d', 'audience',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('Orders')),
            _clean_int(r.get('Units sold')),
            _clean_num(r.get('Sales')),
            None,  # acos
            _clean_num(r.get('ROAS')),
            None,  # cvr
            None,  # top-of-search
            None, None, None, None,  # own/other sku
            None, None, None,  # ntb, branded
            now,
        ))

    # SD Matched Target (2 rows, fold into targeting)
    for r in _read_report(ADS_DIR / 'sd_matched_target.xlsx'):
        month = _extract_month(r)
        all_rows.append((
            month, 'SD',
            (r.get('Campaign Name') or '').strip(),
            '',
            (r.get('Matched target') or r.get('Targeting') or '').strip(),
            '',
            '14d', 'audience',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Total advertiser cost')),
            _clean_int(r.get('14 Day Total Orders (#)')),
            _clean_int(r.get('14 Day Total Units (#)')),
            _clean_num(r.get('14 Day Total Sales')),
            None,  # acos
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            None, None,
            None, None, None, None,
            None, None, None,
            now,
        ))

    _upsert(conn, sql, all_rows)
    print(f"  [ads_targeting] {len(all_rows)} rows")
    _log_pull(conn, 'ads_targeting', '', 'sp+sb+sd_targeting', len(all_rows), now)


def load_ads_products(conn) -> None:
    """SP + SD Advertised Product → ads_products."""
    now = _now()
    sql = """
        INSERT INTO ads_products VALUES %s
        ON CONFLICT (month, ad_type, campaign_name, ad_group_name, asin) DO UPDATE SET
            sku=EXCLUDED.sku, attribution_window=EXCLUDED.attribution_window,
            impressions=EXCLUDED.impressions, clicks=EXCLUDED.clicks,
            ctr=EXCLUDED.ctr, cpc=EXCLUDED.cpc, spend=EXCLUDED.spend,
            orders=EXCLUDED.orders, units=EXCLUDED.units, sales=EXCLUDED.sales,
            acos=EXCLUDED.acos, roas=EXCLUDED.roas, cvr=EXCLUDED.cvr,
            own_sku_units=EXCLUDED.own_sku_units, own_sku_sales=EXCLUDED.own_sku_sales,
            other_sku_units=EXCLUDED.other_sku_units, other_sku_sales=EXCLUDED.other_sku_sales,
            ntb_orders=EXCLUDED.ntb_orders, ntb_sales=EXCLUDED.ntb_sales,
            ntb_units=EXCLUDED.ntb_units, dpv=EXCLUDED.dpv,
            pulled_at=EXCLUDED.pulled_at
    """
    all_rows = []

    # SP Advertised Product
    for r in _read_report(ADS_DIR / 'sp_advertised_product.xlsx'):
        month = _extract_month(r)
        all_rows.append((
            month, 'SP',
            (r.get('Campaign Name') or '').strip(),
            (r.get('Ad Group Name') or '').strip(),
            (r.get('Advertised ASIN') or '').strip(),
            (r.get('Advertised SKU') or '').strip(),
            '7d',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('7 Day Total Orders (#)')),
            _clean_int(r.get('7 Day Total Units (#)')),
            _clean_num(r.get('7 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            _clean_num(r.get('7 Day Conversion Rate')),
            _clean_int(r.get('7 Day Advertised SKU Units (#)')),
            _clean_num(r.get('7 Day Advertised SKU Sales')),
            _clean_int(r.get('7 Day Other SKU Units (#)')),
            _clean_num(r.get('7 Day Other SKU Sales')),
            None, None, None,  # ntb
            None,  # dpv
            now,
        ))

    # SD Advertised Product
    for r in _read_report(ADS_DIR / 'sd_advertised_product.xlsx'):
        month = _extract_month(r)
        all_rows.append((
            month, 'SD',
            (r.get('Campaign Name') or '').strip(),
            (r.get('Ad Group Name') or '').strip(),
            (r.get('Advertised ASIN') or '').strip(),
            (r.get('Advertised SKU') or '').strip(),
            '14d',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('14 Day Total Orders (#)')),
            _clean_int(r.get('14 Day Total Units (#)')),
            _clean_num(r.get('14 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            None,  # cvr
            None, None, None, None,  # own/other sku
            _clean_int(r.get('14 Day New-to-brand Orders (#)')),
            _clean_num(r.get('14 Day New-to-brand Sales')),
            _clean_int(r.get('14 Day New-to-brand Units (#)')),
            _clean_int(r.get('14 Day Detail Page Views (DPV)')),
            now,
        ))

    _upsert(conn, sql, all_rows)
    print(f"  [ads_products] {len(all_rows)} rows")
    _log_pull(conn, 'ads_products', '', 'sp+sd_advertised_product', len(all_rows), now)


def load_ads_placements(conn) -> None:
    """SP Placement + SB Campaign Placement + SB Keyword Placement."""
    now = _now()
    sql = """
        INSERT INTO ads_placements VALUES %s
        ON CONFLICT (month, ad_type, campaign_name, placement,
                     targeting_text, match_type) DO UPDATE SET
            attribution_window=EXCLUDED.attribution_window,
            impressions=EXCLUDED.impressions, clicks=EXCLUDED.clicks,
            ctr=EXCLUDED.ctr, cpc=EXCLUDED.cpc, spend=EXCLUDED.spend,
            orders=EXCLUDED.orders, units=EXCLUDED.units, sales=EXCLUDED.sales,
            acos=EXCLUDED.acos, roas=EXCLUDED.roas,
            ntb_orders=EXCLUDED.ntb_orders, ntb_sales=EXCLUDED.ntb_sales,
            branded_searches=EXCLUDED.branded_searches,
            video_complete_views=EXCLUDED.video_complete_views,
            pulled_at=EXCLUDED.pulled_at
    """
    all_rows = []

    # SP Placement
    for r in _read_report(ADS_DIR / 'sp_placement.xlsx'):
        month = _extract_month(r)
        all_rows.append((
            month, 'SP',
            (r.get('Campaign Name') or '').strip(),
            (r.get('Placement') or '').strip(),
            '', '',  # no targeting/match
            '7d',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            None,  # ctr
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('7 Day Total Orders (#)')),
            _clean_int(r.get('7 Day Total Units (#)')),
            _clean_num(r.get('7 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            None, None, None, None,  # ntb, branded, video
            now,
        ))

    # SB Campaign Placement
    for r in _read_report(ADS_DIR / 'sb_campaign_placement.xlsx'):
        month = _extract_month(r)
        all_rows.append((
            month, 'SB',
            (r.get('Campaign Name') or '').strip(),
            (r.get('Placement') or '').strip(),
            '', '',
            '14d',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('14 Day Total Orders (#)')),
            _clean_int(r.get('14 Day Total Units (#)')),
            _clean_num(r.get('14 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            _clean_int(r.get('14 Day New-to-brand Orders (#)')),
            _clean_num(r.get('14 Day New-to-brand Sales')),
            _clean_int(r.get('14 Day Branded Searches')),
            _clean_int(r.get('Video Complete Views')),
            now,
        ))

    # SB Keyword Placement — skip rows with empty placement (aggregate rows)
    for r in _read_report(ADS_DIR / 'sb_keyword_placement.xlsx'):
        placement = (r.get('Placement Type') or '').strip()
        if not placement:
            continue
        month = _extract_month(r)
        all_rows.append((
            month, 'SB',
            (r.get('Campaign Name') or '').strip(),
            placement,
            (r.get('Targeting') or '').strip(),
            (r.get('Match Type') or '').strip(),
            '14d',
            _clean_int(r.get('Impressions')),
            _clean_int(r.get('Clicks')),
            _clean_num(r.get('Click-Thru Rate (CTR)')),
            _clean_num(r.get('Cost Per Click (CPC)')),
            _clean_num(r.get('Spend')),
            _clean_int(r.get('14 Day Total Orders (#)')),
            _clean_int(r.get('14 Day Total Units (#)')),
            _clean_num(r.get('14 Day Total Sales')),
            _clean_num(r.get('Total Advertising Cost of Sales (ACOS)')),
            _clean_num(r.get('Total Return on Advertising Spend (ROAS)')),
            _clean_int(r.get('14 Day New-to-brand Orders (#)')),
            _clean_num(r.get('14 Day New-to-brand Sales')),
            None, None,  # branded, video
            now,
        ))

    # Deduplicate by PK (month, ad_type, campaign_name, placement, targeting_text, match_type)
    seen: dict[tuple, tuple] = {}
    for row in all_rows:
        pk = (row[0], row[1], row[2], row[3], row[4], row[5])
        seen[pk] = row  # last wins
    deduped = list(seen.values())

    _upsert(conn, sql, deduped)
    print(f"  [ads_placements] {len(deduped)} rows (from {len(all_rows)} raw)")
    _log_pull(conn, 'ads_placements', '', 'sp+sb_placement', len(deduped), now)


def load_ads_purchased_products(conn) -> None:
    """SP Purchased Product → ads_purchased_products."""
    now = _now()
    sql = """
        INSERT INTO ads_purchased_products VALUES %s
        ON CONFLICT (month, campaign_name, advertised_asin,
                     purchased_asin, targeting_text) DO UPDATE SET
            ad_group_name=EXCLUDED.ad_group_name,
            match_type=EXCLUDED.match_type,
            other_sku_units=EXCLUDED.other_sku_units,
            other_sku_orders=EXCLUDED.other_sku_orders,
            other_sku_sales=EXCLUDED.other_sku_sales,
            pulled_at=EXCLUDED.pulled_at
    """
    rows = []
    for r in _read_report(ADS_DIR / 'sp_purchased_product.xlsx'):
        month = _extract_month(r)
        rows.append((
            month,
            (r.get('Campaign Name') or '').strip(),
            (r.get('Ad Group Name') or '').strip(),
            (r.get('Advertised ASIN') or '').strip(),
            (r.get('Purchased ASIN') or '').strip(),
            (r.get('Targeting') or '').strip(),
            (r.get('Match Type') or '').strip(),
            _clean_int(r.get('7 Day Other SKU Units (#)')),
            _clean_int(r.get('7 Day Other SKU Orders (#)')),
            _clean_num(r.get('7 Day Other SKU Sales')),
            now,
        ))
    _upsert(conn, sql, rows)
    print(f"  [ads_purchased_products] {len(rows)} rows")
    _log_pull(conn, 'ads_purchased_products', '', 'sp_purchased_product', len(rows), now)


def load_ads_benchmarks(conn) -> None:
    """Cross-channel Benchmarks + SB Category Benchmark → ads_benchmarks."""
    now = _now()
    sql = """
        INSERT INTO ads_benchmarks VALUES %s
        ON CONFLICT (month, ad_type, category, brand, metric_name) DO UPDATE SET
            your_value=EXCLUDED.your_value, p25=EXCLUDED.p25,
            p50=EXCLUDED.p50, p75=EXCLUDED.p75,
            peer_set_size=EXCLUDED.peer_set_size,
            pulled_at=EXCLUDED.pulled_at
    """
    all_rows = []

    # Cross-channel benchmarks
    metric_cols = [
        ('Percent of purchases new to brand', 'ntb_purchase_pct'),
        ('Purchase rate (new to brand)', 'ntb_purchase_rate'),
        ('Cost per purchase (new to brand)', 'ntb_cost_per_purchase'),
        ('CTR', 'ctr'),
        ('CPC', 'cpc'),
        ('CPM', 'cpm'),
        ('Completion rate (video ad)', 'video_completion_rate'),
        ('Cost per completed view (video ad)', 'video_cpv'),
    ]
    for r in _read_report(ADS_DIR / 'benchmarks.xlsx'):
        month_str = str(r.get('Date', ''))
        import pandas as pd
        try:
            month = pd.to_datetime(month_str).strftime('%Y-%m')
        except Exception:
            month = '2026-02'
        ad_type = (r.get('Ad product') or '').strip()
        category = (r.get('Browse category') or '').strip()
        brand = (r.get('Brand') or '').strip()
        for col_prefix, metric_name in metric_cols:
            val = _clean_num(r.get(col_prefix))
            if val is None:
                continue
            all_rows.append((
                month, ad_type, category, brand, metric_name,
                val,
                _clean_num(r.get(f'{col_prefix} (p25)')),
                _clean_num(r.get(f'{col_prefix} (p50)')),
                _clean_num(r.get(f'{col_prefix} (p75)')),
                _clean_int(r.get('Peer set size')),
                now,
            ))

    # SB Category Benchmark
    sb_metric_cols = [
        ('Click-Thru Rate (CTR)', 'Peer CTR', 'ctr'),
        ('Impressions', 'Peer impressions', 'impressions'),
        ('Total Advertising Cost of Sales (ACoS)', 'Peer ACOS', 'acos'),
        ('Total Return on Advertising Spend (RoAS)', 'Peer ROAS', 'roas'),
    ]
    for r in _read_report(ADS_DIR / 'sb_category_benchmark.csv'):
        month = _extract_month(r)
        category = (r.get('Category') or '').strip()
        brand = (r.get('Brand') or '').strip()
        for val_col, peer_prefix, metric_name in sb_metric_cols:
            val = _clean_num(r.get(val_col))
            if val is None:
                continue
            all_rows.append((
                month, 'SB', category, brand, metric_name,
                val,
                _clean_num(r.get(f'{peer_prefix} - bottom 25%')),
                _clean_num(r.get(f'{peer_prefix} - median')),
                _clean_num(r.get(f'{peer_prefix} - top 25%')),
                None,
                now,
            ))

    # Deduplicate by PK (month, ad_type, category, brand, metric_name)
    seen: dict[tuple, tuple] = {}
    for row in all_rows:
        pk = (row[0], row[1], row[2], row[3], row[4])
        seen[pk] = row
    deduped = list(seen.values())

    _upsert(conn, sql, deduped)
    print(f"  [ads_benchmarks] {len(deduped)} rows (from {len(all_rows)} raw)")
    _log_pull(conn, 'ads_benchmarks', '', 'benchmarks+sb_category', len(deduped), now)


def load_ads_invalid_traffic(conn) -> None:
    """SP + SB + SD Gross/Invalid Traffic → ads_invalid_traffic."""
    now = _now()
    sql = """
        INSERT INTO ads_invalid_traffic VALUES %s
        ON CONFLICT (month, ad_type, campaign_name) DO UPDATE SET
            gross_impressions=EXCLUDED.gross_impressions,
            impressions=EXCLUDED.impressions,
            invalid_impressions=EXCLUDED.invalid_impressions,
            invalid_impression_rate=EXCLUDED.invalid_impression_rate,
            gross_clicks=EXCLUDED.gross_clicks, clicks=EXCLUDED.clicks,
            invalid_clicks=EXCLUDED.invalid_clicks,
            invalid_click_rate=EXCLUDED.invalid_click_rate,
            pulled_at=EXCLUDED.pulled_at
    """
    all_rows = []
    for ad_type, fname in [('SP', 'sp_invalid_traffic.xlsx'),
                           ('SB', 'sb_invalid_traffic.xlsx'),
                           ('SD', 'sd_invalid_traffic.xlsx')]:
        for r in _read_report(ADS_DIR / fname):
            month = _extract_month(r)
            all_rows.append((
                month, ad_type,
                (r.get('Campaign Name') or '').strip(),
                _clean_int(r.get('Gross Impressions')),
                _clean_int(r.get('Impressions')),
                _clean_int(r.get('Invalid Impressions')),
                _clean_num(r.get('Invalid Impression Rate')),
                _clean_int(r.get('Gross Clicks')),
                _clean_int(r.get('Clicks')),
                _clean_int(r.get('Invalid Clicks')),
                _clean_num(r.get('Invalid Click Rate')),
                now,
            ))
    _upsert(conn, sql, all_rows)
    print(f"  [ads_invalid_traffic] {len(all_rows)} rows")
    _log_pull(conn, 'ads_invalid_traffic', '', 'sp+sb+sd_invalid_traffic', len(all_rows), now)


# ---------------------------------------------------------------------------
# Loader map & CLI
# ---------------------------------------------------------------------------

ADS_LOADER_MAP = {
    "campaigns":          ("Ad Campaigns",          load_ads_campaigns),
    "search_terms":       ("Ad Search Terms",       load_ads_search_terms),
    "targeting":          ("Ad Targeting",           load_ads_targeting),
    "products":           ("Ad Products",            load_ads_products),
    "placements":         ("Ad Placements",          load_ads_placements),
    "purchased_products": ("Ad Purchased Products",  load_ads_purchased_products),
    "benchmarks":         ("Ad Benchmarks",          load_ads_benchmarks),
    "invalid_traffic":    ("Ad Invalid Traffic",     load_ads_invalid_traffic),
}


def load_all(conn, only: str | None = None) -> None:
    if only:
        if only not in ADS_LOADER_MAP:
            print(f"Unknown --only value: {only}")
            print(f"Valid: {', '.join(ADS_LOADER_MAP.keys())}")
            return
        label, fn = ADS_LOADER_MAP[only]
        print(f"Loading {label}...")
        fn(conn)
        return

    for key, (label, fn) in ADS_LOADER_MAP.items():
        print(f"Loading {label}...")
        fn(conn)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Load Amazon Advertising data")
    parser.add_argument("--only", choices=list(ADS_LOADER_MAP.keys()),
                        help="Load only this report type")
    args = parser.parse_args()

    init_db()
    conn = get_conn()
    try:
        load_all(conn, only=args.only)
    finally:
        conn.close()
    print("Done.")
